import serial
import time
from collections import deque
import os
import shutil
from openpyxl import load_workbook

# -----------------------------
# USER SETTINGS
# -----------------------------
PORT = "COM6"
BAUDRATE = 115200

TEMPLATE_FILE = "Prusa_Gcode/data/encoder_template.xlsm"
OUTPUT_FILE = "../ME/calibration/batches/testi/encoder_calibration.xlsm"

POS_TOL = 0.5
STABILITY_TOL = 0.5
STABILITY_SAMPLES = 10

X_MAX = 32
Y_MAX = 32
Z_MAX = 13
TOTAL_MAX = X_MAX + Y_MAX + Z_MAX  # 77


def get_unique_filename(filepath):
    base, ext = os.path.splitext(filepath)
    counter = 1
    newpath = filepath
    while os.path.exists(newpath):
        newpath = f"{base}_{counter}{ext}"
        counter += 1
    return newpath


# -----------------------------
# SERIAL SETUP (PROPER RESET HANDLING)
# -----------------------------
print("Opening serial connection...")

ser = serial.Serial(PORT, BAUDRATE, timeout=1)

# Proper Arduino reset using DTR toggle
ser.setDTR(False)
time.sleep(1)
ser.reset_input_buffer()
ser.setDTR(True)

print("Waiting for Arduino to begin streaming data...")

# Wait until first valid numeric CSV line appears
while True:
    line = ser.readline().decode("utf-8", errors="ignore").strip()
    if not line:
        continue

    parts = line.split(",")
    if len(parts) == 7:
        try:
            list(map(float, parts))
            print("Arduino ready.")
            break
        except:
            continue

print("Waiting for printer to reach X0 Y0...")


# -----------------------------
# STATE
# -----------------------------
logging_armed = False
motion_detected = False
test_completed = False

x_buf = deque(maxlen=STABILITY_SAMPLES)
y_buf = deque(maxlen=STABILITY_SAMPLES)
z_buf = deque(maxlen=STABILITY_SAMPLES)

x_count = 0
y_count = 0
z_count = 0


# -----------------------------
# LOGGING LOOP
# -----------------------------
while True:
    try:
        if (x_count + y_count + z_count) >= TOTAL_MAX:
            print("All 77 points captured. Logging complete.")
            test_completed = True
            break

        line = ser.readline().decode("utf-8", errors="ignore").strip()
        if not line:
            continue

        parts = line.split(",")

        if len(parts) != 7:
            continue

        try:
            millis, x, y, z, dx, dy, dz = map(float, parts)
        except:
            continue

        # -----------------------------
        # ARM LOGGING AT X0 Y0
        # -----------------------------
        if not logging_armed:
            if abs(x) <= POS_TOL and abs(y) <= POS_TOL:
                logging_armed = True
                print("Reached X0 Y0 — logging armed")
            else:
                continue

        # -----------------------------
        # MOTION DETECTION
        # -----------------------------
        moving = abs(dx) > 0.01 or abs(dy) > 0.01 or abs(dz) > 0.01

        if moving:
            motion_detected = True
            x_buf.clear()
            y_buf.clear()
            z_buf.clear()
            continue

        if not motion_detected:
            continue

        # -----------------------------
        # SELECT ACTIVE AXIS
        # -----------------------------
        if x_count < X_MAX:
            axis = "X"
            buf = x_buf
            value = x
        elif y_count < Y_MAX:
            axis = "Y"
            buf = y_buf
            value = y
        elif z_count < Z_MAX:
            axis = "Z"
            buf = z_buf
            value = z
        else:
            continue

        buf.append(value)

        # -----------------------------
        # STABILITY CHECK
        # -----------------------------
        if len(buf) < STABILITY_SAMPLES:
            continue

        if max(buf) - min(buf) > STABILITY_TOL:
            continue

        # -----------------------------
        # LOG ONE VALUE ONLY
        # -----------------------------
        index = x_count + y_count + z_count

        # Store data temporarily in memory
        if 'logged_data' not in globals():
            logged_data = []

        logged_data.append([index, axis, value])

        if axis == "X":
            x_count += 1
        elif axis == "Y":
            y_count += 1
        else:
            z_count += 1

        print(f"Logged {axis} point {index}: {value}")

        motion_detected = False
        x_buf.clear()
        y_buf.clear()
        z_buf.clear()

    except KeyboardInterrupt:
        print("\nLogging stopped by user.")
        break

    except Exception as e:
        print("Runtime error:", e)


# -----------------------------
# CLEANUP
# -----------------------------
ser.close()


# -----------------------------
# SAVE ONLY IF COMPLETE
# -----------------------------
if test_completed:

    OUTPUT_FILE = get_unique_filename(OUTPUT_FILE)
    shutil.copyfile(TEMPLATE_FILE, OUTPUT_FILE)

    wb = load_workbook(OUTPUT_FILE, keep_vba=True)

    if "RawData" in wb.sheetnames:
        ws = wb["RawData"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("RawData")

    for sheet in wb.sheetnames:
        if sheet != "RawData":
            del wb[sheet]

    ws.append(["index", "axis", "value"])

    for row in logged_data:
        ws.append(row)

    wb.save(OUTPUT_FILE)
    wb.close()

    print(f"Data saved to {OUTPUT_FILE}")

else:
    print("Test did not complete — no file created.")