import datetime
import json
import os
import shutil
import stat
import subprocess
import sys
import tempfile
import time
import zipfile

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from final_codes.sequential_generate_stt import generate_pydeck
from final_codes.sequential_consolidation import MAX_TEMP_FILTER, MIN_TEMP_FILTER, merge_encoder_camera
from final_codes.sequential_encoder import EncoderLogger
from final_codes.sequential_camera import IRCameraLogger


try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ---------------------------------
# Configuration
# ---------------------------------
SPECIMEN_ID = "S4"
POS_A = (30.0, 30.0, 2.0)
POS_X = POS_A
POS_B = (30.0, 30.0, 100.0)

TOTAL_SUBPRINTS = 5
TOLERANCE = 4  # mm
STABLE_TIME = 5  # seconds


# ---------------------------------
# Utility Functions
# ---------------------------------
def save_metadata(run_folder, ir_camera, encoder, specimen_id, part_id, run_index, total_runs, status="started"):
    """Create or update a JSON metadata file for the current subprint."""

    metadata_path = os.path.join(run_folder, "metadata.json")
    existing_data = {}
    if os.path.exists(metadata_path):
        try:
            with open(metadata_path, "r", encoding="utf-8") as metadata_file:
                existing_data = json.load(metadata_file)
        except (OSError, json.JSONDecodeError):
            existing_data = {}

    data = {
        "session_id": specimen_id,
        "specimen_id": specimen_id,
        "part_id": part_id,
        "subprint_id": os.path.basename(run_folder),
        "status": status,
        "run_index": run_index,
        "total_runs": total_runs,
        "start_time": existing_data.get("start_time") or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "last_updated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "hardware": {
            "encoder_port": "COM6",
            "camera_device": ir_camera.device,
            "emissivity": ir_camera.emissivity,
            "center_point_emissivity": ir_camera.center_point_emissivity,
            "Fan": "Off",
            "Gain Mode": "High",
            "Material": "PA6-CF",
        },
        "parameters": {
            "center_position": ir_camera.center_position,
            "pixel_points": ir_camera.pixel_points,
            "measured_point_count": ir_camera.measured_point_count,
            "pixel_coordinate_mode": "absolute",
            "tolerance_mm": TOLERANCE,
            "Infill speed": "10 mm/s",
            "Min filter": str(MIN_TEMP_FILTER),
            "Max filter": str(MAX_TEMP_FILTER),
            "Extrusion multiplier": "1",
        },
        "files": {
            "encoder_raw": encoder.output_file,
            "encoder_filtered": encoder.filtered_file,
            "camera_csv": ir_camera.csv_filename,
        },
    }

    with open(metadata_path, "w", encoding="utf-8") as metadata_file:
        json.dump(data, metadata_file, indent=4)
    print(f"Metadata saved to {metadata_path}")

DATA_ROOT = os.path.join(PROJECT_ROOT, "ltg_data")
TRACKER_WORKBOOK = os.path.abspath(
    os.path.join(
        PROJECT_ROOT,
        "..",
        "main_study",
        "pa6cf_study",
        "PA6CF_production_workflow_tracker.xlsx",
    )
)
EXPECTED_SPECIMEN_IDS = [f"S{i}" for i in range(1, 6)]
EXPECTED_PART_IDS = [f"P{i}" for i in range(1, 6)]
TRACKER_BACKUP_DIRNAME = "_tracker_backups"


class ProductionWorkflowTracker:
    BATCH_SHEET = "Batch_Tracker"
    PART_SHEET = "Part_Tracker"

    BATCH_REQUIRED_COLUMNS = {
        "Specimen ID": ["Specimen ID"],
        "Batch_folder_path": ["Batch_folder_path"],
        "Print Start": ["Print Start"],
        "Print End": ["Print End"],
    }
    PART_REQUIRED_COLUMNS = {
        "Specimen ID": ["Specimen ID"],
        "Part ID": ["Part ID"],
        "Part_folder_path": ["Part_folder_path", "Part_folder path"],
    }

    def __init__(self, workbook_path):
        self.workbook_path = os.path.abspath(workbook_path)
        self.workspace_root = PROJECT_ROOT
        self._batch_headers = {}
        self._part_headers = {}
        self._batch_rows = {}
        self._part_rows = {}
        self._part_folder_column_index = None

    def validate_for_production(self):
        if load_workbook is None:
            raise RuntimeError(
                "openpyxl is required for safe Excel logging but is not installed. "
                "Install openpyxl before running this workflow."
            )

        if not os.path.isfile(self.workbook_path):
            raise FileNotFoundError(f"Production tracker workbook not found: {self.workbook_path}")

        self._assert_workbook_writable()
        workbook = self._open_workbook()
        try:
            self._validate_sheet_names(workbook)
            self._batch_headers = self._resolve_headers(
                workbook[self.BATCH_SHEET],
                self.BATCH_SHEET,
                self.BATCH_REQUIRED_COLUMNS,
            )
            self._part_headers = self._resolve_headers(
                workbook[self.PART_SHEET],
                self.PART_SHEET,
                self.PART_REQUIRED_COLUMNS,
            )
            self._part_folder_column_index = self._part_headers["Part_folder_path"]
            self._batch_rows = self._index_batch_rows(workbook[self.BATCH_SHEET])
            self._part_rows = self._index_part_rows(workbook[self.PART_SHEET])
            self._validate_expected_rows()
        finally:
            workbook.close()

    def validate_specimen_id(self, specimen_id):
        specimen_id = (specimen_id or "").strip().upper()
        if not specimen_id:
            raise RuntimeError(
                "SPECIMEN_ID is empty. Set SPECIMEN_ID near the top of seq_main_structure.py "
                "before running the production workflow."
            )
        if specimen_id not in EXPECTED_SPECIMEN_IDS:
            raise RuntimeError(
                f"SPECIMEN_ID must be one of {', '.join(EXPECTED_SPECIMEN_IDS)}, got '{specimen_id}'."
            )
        if specimen_id not in self._batch_rows:
            raise RuntimeError(
                f"Specimen '{specimen_id}' is not present in '{self.BATCH_SHEET}'."
            )
        missing_parts = [
            part_id for part_id in EXPECTED_PART_IDS if (specimen_id, part_id) not in self._part_rows
        ]
        if missing_parts:
            raise RuntimeError(
                f"Specimen '{specimen_id}' is missing part rows in '{self.PART_SHEET}': "
                f"{', '.join(missing_parts)}"
            )
        return specimen_id

    def update_batch_folder_path(self, specimen_id, folder_path):
        self._update_cell_value(
            self.BATCH_SHEET,
            self._batch_rows[specimen_id],
            self._batch_headers["Batch_folder_path"],
            self._relative_workspace_path(folder_path),
        )

    def update_batch_print_start(self, specimen_id, timestamp):
        self._update_cell_value(
            self.BATCH_SHEET,
            self._batch_rows[specimen_id],
            self._batch_headers["Print Start"],
            timestamp,
        )

    def update_batch_print_end(self, specimen_id, timestamp):
        self._update_cell_value(
            self.BATCH_SHEET,
            self._batch_rows[specimen_id],
            self._batch_headers["Print End"],
            timestamp,
        )

    def update_part_folder_path(self, specimen_id, part_id, folder_path):
        self._update_cell_value(
            self.PART_SHEET,
            self._part_rows[(specimen_id, part_id)],
            self._part_folder_column_index,
            self._relative_workspace_path(folder_path),
        )

    def _assert_workbook_writable(self):
        try:
            with open(self.workbook_path, "rb+"):
                pass
        except OSError as exc:
            raise RuntimeError(
                f"Production tracker workbook is not writable. Close it in Excel and retry: "
                f"{self.workbook_path}"
            ) from exc

    def _validate_sheet_names(self, workbook):
        actual_sheet_names = set(workbook.sheetnames)
        required_sheet_names = {self.BATCH_SHEET, self.PART_SHEET}
        missing = sorted(required_sheet_names - actual_sheet_names)
        if missing:
            raise RuntimeError(
                f"Production tracker workbook is missing required sheets: {', '.join(missing)}"
            )

    def _resolve_headers(self, worksheet, sheet_name, required_headers):
        header_row = self._row_to_mapping(worksheet, 1)
        resolved = {}

        for logical_name, accepted_names in required_headers.items():
            match = next((name for name in accepted_names if name in header_row), None)
            if match is None:
                accepted_text = ", ".join(accepted_names)
                raise RuntimeError(
                    f"Sheet '{sheet_name}' is missing required column '{logical_name}'. "
                    f"Accepted header names: {accepted_text}"
                )
            resolved[logical_name] = header_row[match]

        return resolved

    def _index_batch_rows(self, worksheet):
        specimen_column = self._batch_headers["Specimen ID"]
        rows = {}
        for row_number in range(2, worksheet.max_row + 1):
            specimen_id = self._normalize_cell_value(worksheet.cell(row=row_number, column=specimen_column).value)
            if not specimen_id:
                continue
            if specimen_id in rows:
                raise RuntimeError(
                    f"Sheet '{self.BATCH_SHEET}' contains duplicate Specimen ID '{specimen_id}'."
                )
            rows[specimen_id] = row_number
        return rows

    def _index_part_rows(self, worksheet):
        specimen_column = self._part_headers["Specimen ID"]
        part_column = self._part_headers["Part ID"]
        rows = {}
        for row_number in range(2, worksheet.max_row + 1):
            specimen_id = self._normalize_cell_value(worksheet.cell(row=row_number, column=specimen_column).value)
            part_id = self._normalize_cell_value(worksheet.cell(row=row_number, column=part_column).value)
            if not specimen_id or not part_id:
                continue
            row_key = (specimen_id, part_id)
            if row_key in rows:
                raise RuntimeError(
                    f"Sheet '{self.PART_SHEET}' contains duplicate row for {specimen_id}/{part_id}."
                )
            rows[row_key] = row_number
        return rows

    def _validate_expected_rows(self):
        missing_batch_rows = [specimen_id for specimen_id in EXPECTED_SPECIMEN_IDS if specimen_id not in self._batch_rows]
        if missing_batch_rows:
            raise RuntimeError(
                f"Sheet '{self.BATCH_SHEET}' is missing specimen rows: {', '.join(missing_batch_rows)}"
            )

        missing_part_rows = [
            f"{specimen_id}/{part_id}"
            for specimen_id in EXPECTED_SPECIMEN_IDS
            for part_id in EXPECTED_PART_IDS
            if (specimen_id, part_id) not in self._part_rows
        ]
        if missing_part_rows:
            raise RuntimeError(
                f"Sheet '{self.PART_SHEET}' is missing part rows: {', '.join(missing_part_rows)}"
            )

    def _row_to_mapping(self, worksheet, row_number):
        mapping = {}
        for column_number, cell in enumerate(worksheet[row_number], start=1):
            cell_value = self._normalize_cell_value(cell.value)
            if cell_value:
                mapping[cell_value] = column_number
        return mapping

    def _update_cell_value(self, sheet_name, row_number, column_number, value):
        workbook = self._open_workbook()
        try:
            worksheet = workbook[sheet_name]
            worksheet.cell(row=row_number, column=column_number).value = value
            self._mark_calculate_on_open(workbook)
            self._save_workbook_safely(workbook)
        finally:
            workbook.close()

    def _open_workbook(self):
        try:
            return load_workbook(self.workbook_path)
        except zipfile.BadZipFile as exc:
            raise RuntimeError(
                f"Production tracker workbook is already corrupted and cannot be opened safely: "
                f"{self.workbook_path}"
            ) from exc

    def _mark_calculate_on_open(self, workbook):
        calculation = getattr(workbook, "calculation", None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True
            calculation.forceFullCalc = True

    def _save_workbook_safely(self, workbook):
        workbook_dir = os.path.dirname(self.workbook_path)
        backup_dir = os.path.join(workbook_dir, TRACKER_BACKUP_DIRNAME)
        os.makedirs(backup_dir, exist_ok=True)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_path = os.path.join(
            backup_dir,
            f"{os.path.splitext(os.path.basename(self.workbook_path))[0]}_{timestamp}.xlsx",
        )
        shutil.copy2(self.workbook_path, backup_path)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=workbook_dir) as temp_file:
            temp_path = temp_file.name

        try:
            workbook.save(temp_path)
            self._verify_workbook_file(temp_path)
            os.replace(temp_path, self.workbook_path)
        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise

    def _verify_workbook_file(self, workbook_path):
        if not zipfile.is_zipfile(workbook_path):
            raise RuntimeError(f"Saved workbook is not a valid .xlsx file: {workbook_path}")

        try:
            verification_workbook = load_workbook(workbook_path)
        except Exception as exc:
            raise RuntimeError(
                f"Saved workbook verification failed. Original workbook was preserved in backup."
            ) from exc
        else:
            verification_workbook.close()

    def _normalize_cell_value(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _relative_workspace_path(self, folder_path):
        absolute_path = os.path.abspath(folder_path)
        try:
            common_root = os.path.commonpath([self.workspace_root, absolute_path])
        except ValueError as exc:
            raise RuntimeError(
                f"Cannot express path relative to workspace root '{self.workspace_root}': {absolute_path}"
            ) from exc

        if common_root != self.workspace_root:
            raise RuntimeError(
                f"Path is outside workspace root '{self.workspace_root}': {absolute_path}"
            )

        return os.path.relpath(absolute_path, self.workspace_root)


def part_id_from_run_index(run_index):
    return f"P{run_index}"


def create_session_folder(specimen_id):
    session_folder = os.path.abspath(os.path.join(DATA_ROOT, specimen_id))
    if os.path.exists(session_folder) and os.listdir(session_folder):
        raise RuntimeError(
            f"Session folder already exists and is not empty: {session_folder}. "
            "Choose a different specimen or clear the existing folder."
        )
    os.makedirs(session_folder, exist_ok=True)
    return session_folder


def create_subprint_folder(session_folder, specimen_id, part_id):
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_folder = os.path.abspath(os.path.join(session_folder, f"{specimen_id}_{timestamp}_{part_id}"))
    os.makedirs(run_folder, exist_ok=False)
    return run_folder



def delete_session_folder(session_folder):
    """
    Force-delete a session folder on Windows or OneDrive-backed storage.
    """
    if not os.path.exists(session_folder):
        return

    def remove_readonly(func, path, excinfo):
        os.chmod(path, stat.S_IWRITE)
        func(path)

    print(f"Attempting to force-delete: {session_folder}")

    for _ in range(3):
        try:
            shutil.rmtree(session_folder, onerror=remove_readonly)
            print("Folder deleted.")
            return
        except Exception:
            time.sleep(1.0)

    try:
        subprocess.run(["cmd", "/c", "rd", "/s", "/q", os.path.abspath(session_folder)], check=True)
        print("Folder force-deleted via system shell.")
    except Exception as exc:
        print(f"System shell could not delete the folder: {exc}")


def within_tolerance(current, target, tol):
    return all(abs(c - t) <= tol for c, t in zip(current, target))


def within_volume(current, bounds):
    x, y, z = current
    xmin, xmax, ymin, ymax, zmin, zmax = bounds
    return xmin <= x <= xmax and ymin <= y <= ymax and zmin <= z <= zmax


def wait_until_stable(get_position_func, target, abort_check, ui_tick=None):
    """
    Wait until position is within tolerance and remains stable for STABLE_TIME seconds.
    """
    print(f"Waiting for stable position at {target} ...")
    stable_start = None

    while True:
        if ui_tick is not None:
            ui_tick()

        if abort_check():
            return False

        pos = get_position_func()

        if pos is None:
            time.sleep(0.01)
            continue

        if within_tolerance(pos, target, TOLERANCE):
            if stable_start is None:
                stable_start = time.time()
            elif time.time() - stable_start >= STABLE_TIME:
                print(f"Stable at {target} for {STABLE_TIME} seconds.")
                return True
        else:
            stable_start = None

        time.sleep(0.01)


def wait_for_manual_encoder_init(ir_camera, abort_check, ui_tick=None):
    print("Camera recording is active.")
    print("Press 'I' in the camera window when you want to initialize the encoder.")

    while not ir_camera.manual_encoder_init_requested:
        if ui_tick is not None:
            ui_tick()

        if abort_check():
            return False
        time.sleep(0.01)

    return True


def wait_for_final_shutdown(get_position_func, target, abort_check, manual_finalize_check, ui_tick=None):
    """
    Wait for stable arrival at POS_B, or allow a manual finalize request
    from the IR camera UI to complete the run and trigger post-processing.
    """
    print(f"Waiting for stable position at {target} or manual finalization request ...")
    stable_start = None

    while True:
        if ui_tick is not None:
            ui_tick()

        if abort_check():
            return "abort"

        if manual_finalize_check():
            print("Manual finalization request received. Proceeding as if POS_B was detected.")
            return "manual"

        pos = get_position_func()

        if pos is None:
            time.sleep(0.01)
            continue

        if within_tolerance(pos, target, TOLERANCE):
            if stable_start is None:
                print(f"Entered final shutdown tolerance band at {pos}.")
                stable_start = time.time()
            elif time.time() - stable_start >= STABLE_TIME:
                print(f"Stable at {target} for {STABLE_TIME} seconds.")
                return "position_b"
        else:
            stable_start = None

        time.sleep(0.01)


def finalize_last_run(
    encoder,
    ir,
    run_folder,
    specimen_id,
    part_id,
    run_index,
    total_runs,
    trigger_source,
    tracker,
):
    ir.stop_recording()
    ir.stop_video()
    encoder.stop_logging()
    encoder.shutdown()
    ir.shutdown()
    tracker.update_batch_print_end(specimen_id, datetime.datetime.now())
    save_metadata(run_folder, ir, encoder, specimen_id, part_id, run_index, total_runs, status="completed")
    run_post_processing(encoder, ir, run_folder)
    print(f"\nFINAL SHUTDOWN TRIGGERED BY {trigger_source}\n")


def wait_until_in_volume(get_position_func, bounds, abort_check, ui_tick=None):
    """
    Wait until the printer enters the current subprint volume before arming POS_X.
    This prevents startup moves near POS_X from ending a subprint early.
    """
    xmin, xmax, ymin, ymax, zmin, zmax = bounds
    print(
        "Waiting for printer to enter subprint volume "
        f"X[{xmin:.1f}, {xmax:.1f}] Y[{ymin:.1f}, {ymax:.1f}] Z[{zmin:.1f}, {zmax:.1f}] ..."
    )

    while True:
        if ui_tick is not None:
            ui_tick()

        if abort_check():
            return False

        pos = get_position_func()
        if pos is None:
            time.sleep(0.01)
            continue

        if within_volume(pos, bounds):
            print(f"Printer entered subprint volume at {pos}.")
            return True

        time.sleep(0.01)


def wait_for_encoder_ready(encoder, abort_check, ui_tick=None):
    print("Waiting for first valid encoder reading...")
    while encoder.get_current_position() is None:
        if ui_tick is not None:
            ui_tick()

        if abort_check():
            return False
        time.sleep(0.01)

    return True


def run_post_processing(encoder, ir, session_folder):
    print("Starting post-processing...")

    encoder_file = encoder.filtered_file
    camera_file = ir.csv_filename
    merged_file = os.path.join(session_folder, "merged_data.csv")
    summary_file = os.path.join(session_folder, "summary_filter.csv")
    html_output = os.path.join(session_folder, "visualization.html")

    merge_encoder_camera(encoder_file, camera_file, merged_file, summary_file)
    print("Merge complete")

    generate_pydeck(merged_file, html_output)
    print("PyDeck visualization created")


def execute_abort_sequence(encoder, ir, session_folder):
    """Kill the current run and delete the full sequential session."""
    print("Abort detected. Signalling threads to stop...")

    if encoder is not None:
        encoder.force_abort = True
        encoder.shutdown()

    if ir is not None:
        ir.force_abort = True
        ir.shutdown()

    delete_session_folder(session_folder)


def finalize_run(encoder, ir, run_folder, specimen_id, part_id, run_index, total_runs, status="completed"):
    ir.stop_recording()
    ir.stop_video()
    encoder.stop_logging()
    encoder.shutdown()
    ir.shutdown()
    save_metadata(run_folder, ir, encoder, specimen_id, part_id, run_index, total_runs, status=status)
    run_post_processing(encoder, ir, run_folder)


# ---------------------------------
# Main Controller
# ---------------------------------
def main():
    tracker = ProductionWorkflowTracker(TRACKER_WORKBOOK)
    tracker.validate_for_production()

    specimen_id = tracker.validate_specimen_id(SPECIMEN_ID)
    session_folder = create_session_folder(specimen_id)
    tracker.update_batch_folder_path(specimen_id, session_folder)

    print(f"\nSequential session folder: {session_folder}\n")
    print(f"Production tracker workbook: {tracker.workbook_path}")
    print(f"Selected specimen: {specimen_id}")
    print("Production tracker preflight passed.\n")

    print("\n=== Sequential Automatic Logging Controller ===")
    print(f"Configured for {TOTAL_SUBPRINTS} subprints.")
    print("Camera recording starts immediately for each subprint.")
    print("Run 1 waits for manual encoder initialization from the camera window.")
    print(f"Run 1 starts at {POS_A}.")
    print(f"Each subprint ends at {POS_X}.")
    print(f"Final shutdown waits for {POS_B}.\n")

    print_started_logged = False

    for run_index in range(1, TOTAL_SUBPRINTS + 1):
        part_id = part_id_from_run_index(run_index)
        run_folder = create_subprint_folder(session_folder, specimen_id, part_id)
        tracker.update_part_folder_path(specimen_id, part_id, run_folder)

        print(f"\n--- Subprint {run_index}/{TOTAL_SUBPRINTS} ({part_id}) ---")
        print(f"Run folder: {run_folder}")

        encoder = EncoderLogger(run_folder, run_index=run_index)
        ir = IRCameraLogger(run_folder, run_index=run_index)

        ir.start()
        ir.start_recording()
        ui_tick = ir.process_ui_events

        check_for_abort = lambda ir=ir: ir.abort_requested
        save_metadata(run_folder, ir, encoder, specimen_id, part_id, run_index, TOTAL_SUBPRINTS, status="recording")

        if run_index == 1:
            if not wait_for_manual_encoder_init(ir, check_for_abort, ui_tick=ui_tick):
                execute_abort_sequence(encoder, ir, session_folder)
                return

            print("Starting encoder initialization...")
            encoder.start()
        else:
            print(f"Subprint {run_index} camera recording is active. Starting encoder immediately.")
            encoder.start()

        if not wait_for_encoder_ready(encoder, check_for_abort, ui_tick=ui_tick):
            execute_abort_sequence(encoder, ir, session_folder)
            return

        ir.mark_encoder_initialized()
        print("Encoder initialized.")

        if run_index == 1:
            if not wait_until_stable(encoder.get_current_position, POS_A, check_for_abort, ui_tick=ui_tick):
                execute_abort_sequence(encoder, ir, session_folder)
                return
            print("\nSTART TRIGGERED AT POS_A\n")
            if not print_started_logged:
                tracker.update_batch_print_start(specimen_id, datetime.datetime.now())
                print_started_logged = True
        else:
            print(f"Subprint {run_index} armed. Waiting for entry into the subprint volume before watching for {POS_X}.")

        encoder.start_logging()
        save_metadata(run_folder, ir, encoder, specimen_id, part_id, run_index, TOTAL_SUBPRINTS, status="recording")

        if not wait_until_in_volume(
            encoder.get_current_position,
            encoder.get_volume_bounds(),
            check_for_abort,
            ui_tick=ui_tick,
        ):
            execute_abort_sequence(encoder, ir, session_folder)
            return

        if run_index < TOTAL_SUBPRINTS:
            if not wait_until_stable(encoder.get_current_position, POS_X, check_for_abort, ui_tick=ui_tick):
                execute_abort_sequence(encoder, ir, session_folder)
                return

            print(f"\nSUBPRINT {run_index} COMPLETE AT POS_X\n")
            finalize_run(encoder, ir, run_folder, specimen_id, part_id, run_index, TOTAL_SUBPRINTS, status="completed")
            print(f"Subprint {run_index} saved. Preparing subprint {run_index + 1}.")
            continue

        save_metadata(
            run_folder,
            ir,
            encoder,
            specimen_id,
            part_id,
            run_index,
            TOTAL_SUBPRINTS,
            status="awaiting_final_shutdown",
        )

        if ir.manual_finalize_requested:
            print("Manual finalization request received before final shutdown wait. Proceeding immediately.")
            final_shutdown_result = "manual"
        else:
            final_shutdown_result = wait_for_final_shutdown(
                encoder.get_current_position,
                POS_B,
                check_for_abort,
                lambda ir=ir: ir.manual_finalize_requested,
                ui_tick=ui_tick,
            )

        if final_shutdown_result == "abort":
            execute_abort_sequence(encoder, ir, session_folder)
            return

        if final_shutdown_result == "manual":
            finalize_last_run(
                encoder,
                ir,
                run_folder,
                specimen_id,
                part_id,
                run_index,
                TOTAL_SUBPRINTS,
                "MANUAL FINALIZATION",
                tracker,
            )
            continue

        finalize_last_run(
            encoder,
            ir,
            run_folder,
            specimen_id,
            part_id,
            run_index,
            TOTAL_SUBPRINTS,
            "POS_B",
            tracker,
        )

    print("Sequential recording completed successfully.")


if __name__ == "__main__":
    main()
